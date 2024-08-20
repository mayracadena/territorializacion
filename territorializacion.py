"""
Para instalar librerias de python en qgis seguir estos pasos:
1. abrir OSGeo4W Shell, en esta consola escribir
2. o4w_env
3. instalar libreria de python con el siguiente codigo 
    3.1. pip install openpyxl
    3.1. pip install os
    3.1. pip install shutil
    
"""

import processing
from qgis.core import QgsVectorLayer, QgsDateTimeFieldFormatter, QgsProject
import openpyxl
import os
import shutil
from PyQt5.QtCore import QDate
from datetime import datetime


class Territorializacion:
    
        
    def manejo_excel(terri, datos):
       
        
        #link del archivo de excel
        directorio_terri = os.path.dirname(__file__)
        archivo_excel_modelo = os.path.join(directorio_terri, 'excel_no_borrar', 'datos_poligonos.xlsx')
        
        archivo_excel = os.path.join(directorio_terri, 'resultado.xlsx')
        
        
        try:
            
            excel_terri = openpyxl.load_workbook(archivo_excel)
        except FileNotFoundError:
            print(f"Error: El archivo {archivo_excel} no fue encontrado.")
            return
            
        hojas = {
            1: 'veredas_DANE',
            2: 'consejos_comunitarios',
            3: 'resguardos_formalizados',
            4: 'solicitudes_ancestrales',
            5: 'solicitudes_coloniales',
            6: 'solicitudes_indigena',
            7: 'solicitudes_negros'
        }   
     
        h_terri = hojas.get(terri, None)
        if not h_terri:
            print(f"Error: No se encontró una hoja para 'terri'={terri}.")
            return

        hoja = excel_terri[h_terri]
        last_row = hoja.max_row
       
        for d in datos:
            last_row = last_row+1
            
            if(terri == 1):
                #veredas_DANE
               for col, value in enumerate(d[:6], start=1):
                    if value is None:
                        return ''
                    elif isinstance(value, QDate):
                        value = value.toString("yyyy-MM-dd")
                    hoja.cell(row=last_row, column=col, value=value)
                    print(value)
            elif(terri == 2 ):
                #consejos_comunitarios
                for col, value in enumerate(d[:14], start=1):
                    if value is None:
                        return ''
                    elif isinstance(value, QDate):
                        value = value.toString("yyyy-MM-dd")
                    hoja.cell(row=last_row, column=col, value=value)
                    print(value)
            elif(terri == 3):
                #resguardos_formalizados
                for col, value in enumerate(d[:12], start=1):
                    if value is None:
                        return ''
                    elif isinstance(value, QDate):
                        value = value.toString("yyyy-MM-dd")
                    hoja.cell(row=last_row, column=col, value=value)
                    print(value)
            elif(terri == 4):
                #solicitudes_ancestrales
                for col, value in enumerate(d[:15], start=1):
                    if value is None:
                        return ''
                    elif isinstance(value, QDate):
                        value = value.toString("yyyy-MM-dd")
                    hoja.cell(row=last_row, column=col, value=value)
                    print(value)
            elif(terri == 5):
                #solicitudes_coloniales
                for col, value in enumerate(d[:15], start=1):
                    if value is None:
                        return ''
                    elif isinstance(value, QDate):
                        value = value.toString("yyyy-MM-dd")
                    hoja.cell(row=last_row, column=col, value=value)
                    print(value)
            elif(terri == 6):
                #solicitudes_indigena
                for col, value in enumerate(d[:15], start=1):
                    if value is None:
                        return ''
                    elif isinstance(value, QDate):
                        value = value.toString("yyyy-MM-dd")
                    hoja.cell(row=last_row, column=col, value=value)
                    print(value)
            
            elif(terri == 7):
                #solicitudes_negros
                for col, value in enumerate(d[:15], start=1):
                    if value is None:
                        return ''
                    elif isinstance(value, QDate):
                        value = value.toString("yyyy-MM-dd")
                        return value
                    hoja.cell(row=last_row, column=col, value=value)
                    print(value)

   
        excel_terri.save(archivo_excel)
        QgsProject.instance().removeMapLayer(iface.activeLayer())
            
     
    def valores_atribututos(datos, terri):
        layer =iface.activeLayer()
        # Get the attribute names
        nombre_atributos = []
        for field in layer.fields():
            nombre_atributos.append(field.name())
            #print(f"se.attribute('{field.name()}'),")
        
        #print(nombre_atributos)
        return nombre_atributos
            
    def capa_activa():
        lyr = iface.activeLayer()
        return lyr
            
    def select_by_location(lyr_input, terri):
        directorio_terri = os.path.dirname(__file__)
        valores = []
        capa = ''
        nombre_capa = ''
        if(terri == 1):
            capa = os.path.join(directorio_terri, 'CAPA VEREDAS DANE 2020', 'CRVeredas_2020','ShapeFile', 'CRVeredas_2020.shp')
            nombre_capa = "CRVeredas_2020"
        elif(terri == 2):
            capa = os.path.join(directorio_terri, 'CAPAS LEGALIZACIÓN', 'FORMALIZADOS ETNICA FEBRERO 12', 'FORMALIZADOS ETNICA FEBRERO 12', 'CONSEJOS COMUNITARIOS.shp' )
            nombre_capa = "CONSEJOS COMUNITARIOS"
        elif(terri == 3):
            capa = os.path.join(directorio_terri,'CAPAS LEGALIZACIÓN', 'FORMALIZADOS ETNICA FEBRERO 12', 'FORMALIZADOS ETNICA FEBRERO 12','RESGUARDOS FORMALIZADOS.shp' )
            nombre_capa = "RESGUARDOS FORMALIZADOS"
        elif(terri == 4):
            capa = os.path.join(directorio_terri, 'CAPAS LEGALIZACIÓN','SOLICITUDES ETNICA FEBRERO12 1','SOLICITUDES ETNICA FEBRERO12','SOLICITUDES COLONIALES.shp')
            nombre_capa = "SOLICITUDES COLONIALES"
        elif(terri == 5):
            capa = os.path.join(directorio_terri,  'CAPAS LEGALIZACIÓN','SOLICITUDES ETNICA FEBRERO12 1','SOLICITUDES ETNICA FEBRERO12','SOLICITUDES INDIGENA.shp')
            nombre_capa = "SOLICITUDES INDIGENAS"
        elif(terri == 6):
            capa = os.path.join(directorio_terri,  'CAPAS LEGALIZACIÓN','SOLICITUDES ETNICA FEBRERO12 1','SOLICITUDES ETNICA FEBRERO12','SOLICITUDES NEGROS.shp')
            nombre_capa = "SOLICITUDES NEGROS"
        elif(terri == 7):
            capa = os.path.join(directorio_terri,  'CAPAS LEGALIZACIÓN','SOLICITUDES ETNICA FEBRERO12 1','SOLICITUDES ETNICA FEBRERO12','SOLICITUDES ANCESTRALES.shp')
            nombre_capa = "SOLICITUDES ANCESTRALES"
        
            
        
        ubicacion = iface.addVectorLayer(capa, nombre_capa, "ogr")
        parametros = {
        "INPUT":ubicacion,
        "PREDICATE":0,
        "INTERSECT":lyr_input,
        "METHOD":0,
        "OUTPUT":None}
        
        processing.run("qgis:selectbylocation", parametros)
        valores_seleccionados = ubicacion.selectedFeatures()
              
        for se in valores_seleccionados:
            if(terri == 1):
                #veredas_dane
                valores.append([
                    se.attribute('OBJECTID'),
                    se.attribute('DPTOMPIO'),
                    se.attribute('CODIGO_VER'),
                    se.attribute('NOM_DEP'),
                    se.attribute('NOMB_MPIO'),
                    se.attribute('NOMBRE_VER'),
                    se.attribute('VIGENCIA'),
                    se.attribute('FUENTE'),
                    se.attribute('DESCRIPCIO'),
                    se.attribute('SEUDONIMOS'),
                    se.attribute('AREA_HA'),
                    se.attribute('COD_DPTO'),
                    se.attribute('OBSERVACIO'),
                    se.attribute('CONSEJE'),
                    se.attribute('ORIG_FID'),
                    se.attribute('SHAPE_Leng'),
                    se.attribute('SHAPE_Area')
                ])
            elif(terri == 2):
                #consejos comunitarios
                valores.append([
                    se.attribute('ID_ANT'),
                    se.attribute('NOMBRE'),
                    se.attribute('NUMERO_PLA'),
                    se.attribute('DEPARTAMEN'),
                    se.attribute('MUNICIPIO'),
                    se.attribute('CODIGO_DAN'),
                    se.attribute('GlobalID'),
                    se.attribute('TIPO_ACTO_'),
                    se.attribute('NUMERO_ACT'),
                    se.attribute('FECHA_ACTO'),
                    se.attribute('AREA_ACTO_'),
                    se.attribute('AREA_RECON'),
                    se.attribute('RANG_TOLER'),
                    se.attribute('RECONSTRUC'),
                    se.attribute('RESPONSABL'),
                    se.attribute('AÑO_RECON')
                ])
            elif(terri == 3):
                #resguardos formalizados
                valores.append([
                    se.attribute('ID_ANT'),
                    se.attribute('NOMBRE'),
                    se.attribute('TIPO_ACTO_'),
                    se.attribute('NUMERO_ACT'),
                    se.attribute('FECHA_ACTO'),
                    se.attribute('AREA_ACTO_'),
                    se.attribute('NUMERO_PLA'),
                    se.attribute('PUEBLO'),
                    se.attribute('DEPARTAMEN'),
                    se.attribute('MUNICIPIO'),
                    se.attribute('CODIGO_DAN'),
                    se.attribute('GlobalID'),
                    se.attribute('RECONSTRUC'),
                    se.attribute('RESPONSABL'),
                    se.attribute('AÑO_RECON'),
                    se.attribute('AREA_RECON'),
                    se.attribute('RANG_TOLER')
                ])
            elif(terri == 4):
                #solicitudes ancestrales
                valores.append([
                    se.attribute('ID_SOLICIT'),
                    se.attribute('ID_ANT'),
                    se.attribute('NOMBRE_COM'),
                    se.attribute('DEPARTAMEN'),
                    se.attribute('MUNICIPIO'),
                    se.attribute('LOCALIZACI'),
                    se.attribute('FECHA_SOLI'),
                    se.attribute('CODIGO_EST'),
                    se.attribute('CODIGO_PRO'),
                    se.attribute('ESCENARIO_'),
                    se.attribute('RESPONSABL'),
                    se.attribute('FECHA_EDIC'),
                    se.attribute('AREA_HECTA'),
                    se.attribute('GlobalID'),
                    se.attribute('NUMERO_AMP')
                ])
            elif(terri == 5):#solitudes coloniales
                valores.append([
                    se.attribute('ID_SOLICIT'),
                    se.attribute('ID_ANT'),
                    se.attribute('NOMBRE_COM'),
                    se.attribute('DEPARTAMEN'),
                    se.attribute('MUNICIPIO'),
                    se.attribute('LOCALIZACI'),
                    se.attribute('FECHA_SOLI'),
                    se.attribute('CODIGO_EST'),
                    se.attribute('CODIGO_PRO'),
                    se.attribute('ESCENARIO_'),
                    se.attribute('RESPONSABL'),
                    se.attribute('FECHA_EDIC'),
                    se.attribute('AREA_HECTA'),
                    se.attribute('GlobalID'),
                    se.attribute('NUMERO_AMP')
                ])
                
            elif(terri == 6):
                #solicitudes indigenas
                valores.append([
                    se.attribute('ID_SOLICIT'),
                    se.attribute('ID_ANT'),
                    se.attribute('NOMBRE_COM'),
                    se.attribute('DEPARTAMEN'),
                    se.attribute('MUNICIPIO'),
                    se.attribute('LOCALIZACI'),
                    se.attribute('FECHA_SOLI'),
                    se.attribute('CODIGO_EST'),
                    se.attribute('CODIGO_PRO'),
                    se.attribute('NUMERO_AMP'),
                    se.attribute('ESCENARIO_'),
                    se.attribute('RESPONSABL'),
                    se.attribute('FECHA_EDIC'),
                    se.attribute('AREA_HECTA'),
                    se.attribute('GlobalID')
                ])
            elif(terri == 7):
                #solicitudes negros
                valores.append([
                    se.attribute('ID_SOLICIT'),
                    se.attribute('ID_ANT'),
                    se.attribute('NOMBRE_COM'),
                    se.attribute('DEPARTAMEN'),
                    se.attribute('MUNICIPIO'),
                    se.attribute('LOCALIZACI'),
                    se.attribute('FECHA_SOLI'),
                    se.attribute('CODIGO_EST'),
                    se.attribute('CODIGO_PRO'),
                    se.attribute('ESCENARIO_'),
                    se.attribute('RESPONSABL'),
                    se.attribute('FECHA_EDIC'),
                    se.attribute('AREA_HECTA'),
                    se.attribute('GlobalID'),
                    se.attribute('NUMERO_AMP')
                ])
            
            
    
        return valores
        


        
directorio_terri = os.path.dirname(__file__)
archivo_excel_modelo = os.path.join(directorio_terri, 'excel_no_borrar', 'datos_poligonos.xlsx')
archivo_excel = os.path.join(directorio_terri, 'resultado.xlsx')
shutil.copy(archivo_excel_modelo, archivo_excel)
#VEREDAS
#OPCIONES INDIVIDUALES
"""
seleccionados = Territorializacion.select_by_location(Territorializacion.capa_activa(), 1)
Territorializacion.manejo_excel(1, seleccionados)
Territorializacion.valores_atribututos(Territorializacion.capa_activa, 1)
"""
#OPCION TODAS LAS CAPAS
for i in range(1, 8):
    seleccionados = Territorializacion.select_by_location(Territorializacion.capa_activa(), i)
    Territorializacion.manejo_excel(i, seleccionados)
    Territorializacion.valores_atribututos(Territorializacion.capa_activa, i)
#mensaje en consola para saber si acabo el proceso
print("se termino el proceso, revisa el excel")
